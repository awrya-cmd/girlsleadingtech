"""
Loads documents from disk and splits them into chunks ready for embedding.
Supports: PDF, DOCX, TXT, MD, CSV, XLSX.
"""

from __future__ import annotations

import csv
import io
from pathlib import Path
from typing import List

from langchain_text_splitters import RecursiveCharacterTextSplitter

from config import settings

# Supported extensions (used by ingest_all.py)


# ---------------------------------------------------------------------------
# Chunk dataclass
# ---------------------------------------------------------------------------

class Chunk:
    def __init__(self, text: str, metadata: dict):
        self.text = text
        self.metadata = metadata

    def __repr__(self):
        return f"Chunk(chars={len(self.text)}, source={self.metadata.get('source')})"


# ---------------------------------------------------------------------------
# Per-format loaders
# ---------------------------------------------------------------------------

def _load_pdf(path: Path) -> str:
    from pypdf import PdfReader
    reader = PdfReader(str(path))
    pages = []
    for i, page in enumerate(reader.pages):
        text = page.extract_text() or ""
        if text.strip():
            pages.append(f"[Page {i + 1}]\n{text}")
    return "\n\n".join(pages)


def _load_docx(path: Path) -> str:
    from docx import Document
    doc = Document(str(path))
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    return "\n\n".join(paragraphs)


def _load_txt_or_md(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="ignore")


def _load_csv(path: Path) -> str:
    lines = []
    with open(path, newline="", encoding="utf-8", errors="ignore") as f:
        reader = csv.reader(f)
        for row in reader:
            lines.append(", ".join(row))
    return "\n".join(lines)


def _load_xlsx(path: Path) -> str:
    import openpyxl
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    sheets = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            rows.append(", ".join(cells))
        sheets.append(f"[Sheet: {sheet}]\n" + "\n".join(rows))
    return "\n\n".join(sheets)


LOADERS = {
    ".pdf": _load_pdf,
    ".docx": _load_docx,
    ".doc": _load_docx,
    ".txt": _load_txt_or_md,
    ".md": _load_txt_or_md,
    ".csv": _load_csv,
    ".xlsx": _load_xlsx,
    ".xls": _load_xlsx,
}


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def load_document(path: Path) -> str:
    """Return full text from a file, or raise ValueError for unsupported types."""
    ext = path.suffix.lower()
    loader = LOADERS.get(ext)
    if loader is None:
        raise ValueError(f"Unsupported file type: {ext}")
    return loader(path)


def split_text(text: str, source: str, extra_metadata: dict | None = None) -> List[Chunk]:
    """Split raw text into overlapping chunks with metadata."""
    splitter = RecursiveCharacterTextSplitter(
        chunk_size=settings.chunk_size,
        chunk_overlap=settings.chunk_overlap,
        separators=["\n\n", "\n", ". ", " ", ""],
    )
    raw_chunks = splitter.split_text(text)
    metadata_base = {"source": source, **(extra_metadata or {})}
    return [
        Chunk(text=chunk, metadata={**metadata_base, "chunk_index": i})
        for i, chunk in enumerate(raw_chunks)
        if chunk.strip()
    ]


def process_file(path: Path, extra_metadata: dict | None = None) -> List[Chunk]:
    """Load a file from disk and return its chunks."""
    text = load_document(path)
    return split_text(text, source=path.name, extra_metadata=extra_metadata)


def process_text(text: str, source: str, extra_metadata: dict | None = None) -> List[Chunk]:
    """Process raw text (e.g. from an API payload) directly."""
    return split_text(text, source=source, extra_metadata=extra_metadata)
